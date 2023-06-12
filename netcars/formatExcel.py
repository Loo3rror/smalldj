import openpyxl
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from PIL import Image as PImage
import uuid
import aiohttp
import aiofiles


def makeExel(data,path):
	read = pandas.DataFrame.from_dict(data)
	read.to_excel(path+'.xlsx')
	wb_obj = openpyxl.load_workbook(path+'.xlsx')
	sheet_obj = wb_obj.active
	sheet_obj.column_dimensions['B'].width = 20
	sheet_obj.column_dimensions['C'].width = 15
	sheet_obj.column_dimensions['D'].width = 15
	sheet_obj.column_dimensions['E'].width = 15
	sheet_obj.column_dimensions['F'].width = 15
	sheet_obj.column_dimensions['G'].width = 15
	sheet_obj.column_dimensions['H'].width = 15
	sheet_obj.column_dimensions['I'].width = 15
	sheet_obj.column_dimensions['K'].width = 20
	sheet_obj["A1"].fill = PatternFill("solid", start_color="9987C7")
	sheet_obj["B1"].fill = PatternFill("solid", start_color="9987C7")
	sheet_obj["C1"].fill = PatternFill("solid", start_color="9987C7")
	sheet_obj["D1"].fill = PatternFill("solid", start_color="9987C7")
	sheet_obj["J1"].fill = PatternFill("solid", start_color="9987C7")
	sheet_obj["K1"].fill = PatternFill("solid", start_color="9987C7")
	sheet_obj["E1"].fill = PatternFill("solid", start_color="674EA7")
	sheet_obj["F1"].fill = PatternFill("solid", start_color="674EA7")
	sheet_obj["G1"].fill = PatternFill("solid", start_color="674EA7")
	sheet_obj["H1"].fill = PatternFill("solid", start_color="674EA7")
	sheet_obj["I1"].fill = PatternFill("solid", start_color="674EA7")
	column_img = sheet_obj['B']

	#link style
	column_link = sheet_obj['C']
	for cell in column_link[1:]:
		if cell.value is None:
			break
		cell.style = "Hyperlink"

	#center values
	for row in range(1,sheet_obj.max_row+1):
		for col in range(1,sheet_obj.max_column+1):
			cell=sheet_obj.cell(row, col)
			cell.alignment = Alignment(horizontal='center', vertical='center')
	wb_obj.save(path+'.xlsx')
	loop = asyncio.new_event_loop()
	asyncio.set_event_loop(loop)
	coroutine = getimg(path+'.xlsx')
	parseinfo = loop.run_until_complete(coroutine)
	loop.close()


#add image to cols
async def getimg(path):
	wb_obj = openpyxl.load_workbook(path)
	tmp_list = []
	sheet_obj = wb_obj.active
	column = sheet_obj['B']
	aiohttp_client = aiohttp.ClientSession()
	sheet_obj.column_dimensions['B'].width = 20
	for cell in column[1:]:
		tasks =await aiohttp_client.get(cell.value)
		image_data =await tasks.read()
		pict = str(uuid.uuid4())
		temp_image_filename = os.getcwd()+"/tmp/"+pict+'.jpg'
		tmp_list.append(temp_image_filename)
		async with aiofiles.open(temp_image_filename, 'wb') as f:
			await f.write(image_data)
		try:
			uncropped_img = PImage.open(temp_image_filename)
			img = uncropped_img.resize((140,120))
			img.save(temp_image_filename)
			ximage = img = Image(temp_image_filename)
			sheet_obj.add_image(img, cell.coordinate)
		except:
			print('Null img')
		sheet_obj.row_dimensions[cell.row].height = 80
		wb_obj.save(path)
	for tmp in tmp_list:
		os.remove(tmp)
	await aiohttp_client.close()